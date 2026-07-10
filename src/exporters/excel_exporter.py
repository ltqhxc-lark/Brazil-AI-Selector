# -*- coding: utf-8 -*-
"""
Brazil-AI-Selector - Excel 财务与选品综合报告物理导出器
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.reports.report_models import BusinessReport
from src.exporters.base_exporter import BaseExporter, ExportResult

class ExcelExporter(BaseExporter):
    """
    Excel 格式物理文件导出器
    利用 openpyxl 引擎，将 BusinessReport 树形 DTO 数据编译转化为符合现代企业审计标准、
    排版优美、具备区块视觉区分、自动列宽与自动换行的财务工作簿。
    """

    def supports(self, format_name: str) -> bool:
        return format_name.lower().strip() in ("xlsx", "xls")

    def export(self, report: BusinessReport, output_path: str) -> ExportResult:
        """
        核心物理导出：使用 openpyxl 动态生成财务审计级 Excel 报表
        """
        # 1. 安全前置校验
        if not self.validate(report):
            return ExportResult(
                success=False,
                file_path="",
                message="Excel 导出失败：BusinessReport DTO 结构缺失或内容不完整。",
                errors=["BusinessReport instance validation failed."]
            )

        try:
            parent_dir = os.path.dirname(os.path.abspath(output_path))
            os.makedirs(parent_dir, exist_ok=True)

            # 2. 实例化 openpyxl 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "巴西电商财务与选品诊断"

            # 显式显示网格线
            ws.views.sheetView[0].showGridLines = True

            # 3. 预设专业企业级主题色与样式 (Steel Blue 企业财务经典)
            font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            font_section = Font(name="Calibri", size=12, bold=True, color="1F497D")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Calibri", size=11, bold=True)
            font_regular = Font(name="Calibri", size=11)
            
            fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # 主标题深蓝
            fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # 副标题淡蓝
            fill_accent = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid") # 斑马条高亮

            # 设置预警色背景
            alert_colors = {
                "red": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),    # 红色亏损高亮
                "yellow": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # 黄色警报高亮
                "green": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # 绿色安全高亮
            }
            alert_fill = alert_colors.get(report.summary.overall_alert_level.lower(), alert_colors["green"])

            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            align_right = Alignment(horizontal="right", vertical="center")

            thin_border_side = Side(style='thin', color='BFBFBF')
            double_border_side = Side(style='double', color='1F497D')
            
            border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            border_summary = Border(top=thin_border_side, bottom=double_border_side) # 财务对账底线双线

            # 4. 写入合并主大标题
            ws.merge_cells("A1:C2")
            ws["A1"] = "Brazil-AI-Selector 财务分析与跨境出海综合诊断报告"
            ws["A1"].font = font_title
            ws["A1"].fill = fill_title
            ws["A1"].alignment = align_center

            # 5. 写入报告基本概要区块
            ws.cell(row=4, column=1, value="【A. 综合诊断概要】").font = font_section
            
            summary_data = [
                ("商品 SKU 编码", report.summary.product_sku),
                ("商品标准名称", report.summary.product_name),
                ("目标评测平台", report.summary.platform_name),
                ("综合安全评级", report.summary.overall_alert_level.upper()),
                ("报告生成时间", report.summary.generated_at)
            ]
            
            curr_row = 5
            for label, val in summary_data:
                ws.cell(row=curr_row, column=1, value=label).font = font_bold
                ws.cell(row=curr_row, column=1).border = border_data
                
                target_cell = ws.cell(row=curr_row, column=2, value=val)
                target_cell.font = font_regular
                target_cell.border = border_data
                
                if label == "综合安全评级":
                    target_cell.fill = alert_fill
                    target_cell.font = font_bold
                    
                ws.cell(row=curr_row, column=3).border = border_data # 空单元补边框保持规整
                curr_row += 1

            # 6. 写入核心损益与财务核算区块 (审计表样式)
            curr_row += 1
            ws.cell(row=curr_row, column=1, value="【B. 单笔交易损益核算审计明细】").font = font_section
            
            curr_row += 1
            # 财务标题行
            ws.cell(row=curr_row, column=1, value="财务科目大类").font = font_header
            ws.cell(row=curr_row, column=1).fill = fill_header
            ws.cell(row=curr_row, column=1).alignment = align_center

            ws.cell(row=curr_row, column=2, value="精算雷亚尔金额").font = font_header
            ws.cell(row=curr_row, column=2).fill = fill_header
            ws.cell(row=curr_row, column=2).alignment = align_center

            ws.cell(row=curr_row, column=3, value="科目成本营收占比及税务/平台规则说明").font = font_header
            ws.cell(row=curr_row, column=3).fill = fill_header
            ws.cell(row=curr_row, column=3).alignment = align_center

            finance_rows = [
                ("1. 销售销售标价 (营收)", report.financial.revenue_brl, "100.00%", "销售营收基准"),
                ("2. 商品采购产品进价 (成本)", -report.financial.purchase_cost_brl, f"-{report.financial.formatted_profit_margin}", "供应链到岸货物成本"),
                ("3. 平台收佣及交易附加 (平台费)", -report.financial.total_fee_brl, f"-{report.platform.formatted_total_fee}", "佣金扣点、固定手续费与风险提存"),
                ("4. 政府销售/流转/进口 (税费)", -report.financial.total_tax_brl, f"-{report.tax.formatted_total_tax}", f"采用 {report.tax.tax_regime_name} 税耗"),
                ("5. 卖家实际分摊支付 (物流运费)", -report.financial.total_shipping_brl, f"-{report.financial.formatted_total_shipping}", "包邮运费分摊成本")
            ]

            for item, val, ratio, desc in finance_rows:
                curr_row += 1
                ws.cell(row=curr_row, column=1, value=item).font = font_bold if "营收" in item else font_regular
                ws.cell(row=curr_row, column=1).border = border_data
                
                c_val = ws.cell(row=curr_row, column=2, value=val)
                c_val.font = font_bold if "营收" in item else font_regular
                c_val.border = border_data
                c_val.alignment = align_right
                c_val.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00' # 设置专业财会本地货币格式

                c_desc = ws.cell(row=curr_row, column=3, value=f"占比: {ratio} | {desc}")
                c_desc.font = font_regular
                c_desc.border = border_data
                c_desc.alignment = align_left

            # 纯利润总计结算行
            curr_row += 1
            ws.cell(row=curr_row, column=1, value="单笔成交预估净利润总计").font = font_bold
            ws.cell(row=curr_row, column=1).fill = fill_accent
            ws.cell(row=curr_row, column=1).border = border_summary

            c_prof = ws.cell(row=curr_row, column=2, value=report.financial.net_profit_brl)
            c_prof.font = font_bold
            c_prof.fill = fill_accent
            c_prof.border = border_summary
            c_prof.alignment = align_right
            c_prof.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'

            c_margin = ws.cell(row=curr_row, column=3, value=f"综合纯利润率: {report.financial.formatted_profit_margin} ({report.warnings.alert_description})")
            c_margin.font = font_bold
            c_margin.fill = alert_fill if report.financial.net_profit_brl > 0 else alert_colors["red"]
            c_margin.border = border_summary
            c_margin.alignment = align_left

            # 7. 写入运营、定价、营销与供应链决策区块
            curr_row += 2
            ws.cell(row=curr_row, column=1, value="【C. 策略中心黄金运营决策建议】").font = font_section
            
            strategy_blocks = [
                ("1. 定价调节建议", f"绝对保本底线售价: {report.pricing.formatted_breakeven_price}\n15%黄金净利售价: {report.pricing.formatted_price_at_15_margin}\n20%金牌净利售价: {report.pricing.formatted_price_at_20_margin}\n大促让利安全预算: {report.pricing.formatted_markdown_budget}"),
                ("2. 广告投放与促销", f"推广策略: {report.marketing.marketing_decision_text}\n建议广告费占比控制: {report.marketing.formatted_ads_budget_ratio} 内\n直降优惠券上限额度: {report.marketing.formatted_coupon_cap}"),
                ("3. 供应链渠道备货", f"通关进口链路: {report.inventory.recommended_import_channel}\n补货物流时效: {report.inventory.replenishment_lead_time_days} 天\n建议起订MOQ: {report.inventory.recommended_moq} 件\n安全备货库存: {report.inventory.recommended_safety_stock} 件\n备货核心细节: {report.inventory.inventory_decision_text}")
            ]

            for label, content in strategy_blocks:
                curr_row += 1
                ws.cell(row=curr_row, column=1, value=label).font = font_bold
                ws.cell(row=curr_row, column=1).border = border_data
                ws.cell(row=curr_row, column=1).alignment = align_center
                
                # 合并 B 与 C 列，承载较长的决策文字，开启自动换行
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
                target = ws.cell(row=curr_row, column=2, value=content)
                target.font = font_regular
                target.alignment = align_left
                
                # 给合并出来的 C 列单元格画上框线以保证表格统一
                ws.cell(row=curr_row, column=3).border = border_data
                ws.cell(row=curr_row, column=2).border = border_data

            # 8. 写入行动步骤清单 To-do Action Items
            curr_row += 2
            ws.cell(row=curr_row, column=1, value="【D. 综合落地行动指令清单 (To-do List)】").font = font_section
            
            for idx, action in enumerate(report.recommendations.action_items_list, 1):
                curr_row += 1
                ws.cell(row=curr_row, column=1, value=f"步骤 {idx}").font = font_bold
                ws.cell(row=curr_row, column=1).border = border_data
                ws.cell(row=curr_row, column=1).alignment = align_center
                
                ws.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=3)
                target = ws.cell(row=curr_row, column=2, value=action)
                target.font = font_regular
                target.alignment = align_left
                
                ws.cell(row=curr_row, column=3).border = border_data
                ws.cell(row=curr_row, column=2).border = border_data

            # 9. 智能调整列宽 (A、B列自适应宽度，C列为长文本 wrap 列直接固定大宽度)
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 75

            # 10. 保存工作簿
            wb.save(output_path)

            return ExportResult(
                success=True,
                file_path=os.path.abspath(output_path),
                message=f"成功将财务诊断与运营分析白皮书导出为精美 Excel 工作表！路径: {output_path}"
            )
        except Exception as e:
            return ExportResult(
                success=False,
                file_path="",
                message=f"Excel 导出运行异常: {str(e)}",
                errors=[f"File I/O or openpyxl write error: {str(e)}"]
            )
